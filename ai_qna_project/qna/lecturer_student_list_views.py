# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
import os
import re
import unicodedata
import logging
from datetime import date, datetime
from uuid import uuid4

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl import load_workbook

from .academic_year import (
    ACADEMIC_YEAR_ERROR_MESSAGE,
    is_valid_academic_year,
    sanitize_academic_year_for_display,
)
from .models import (
    SemesterChoices,
    StudentListUploadStatus,
    StudentRosterAccountStatus,
    StudentRosterStudent,
    StudentRosterUpload,
    Subject,
    UserProfile,
)
from .permissions import (
    lecturer_required,
    get_lecturer_subject_or_404,
    lecturer_can_access_subject,
)
from .lecturer_student_accounts import normalize_student_code

User = get_user_model()
LIST_PAGE_SIZE = 8
logger = logging.getLogger(__name__)

def _is_ajax_request(request):
    return (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or "application/json" in request.headers.get("accept", "")
    )

# =========================================================
# STUDENT LIST MANAGEMENT
# =========================================================
STUDENT_LIST_ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
STUDENT_LIST_MAX_FILE_SIZE = 10 * 1024 * 1024
STUDENT_LIST_REQUIRED_COLUMNS = {
    "student_code": {"mssv", "ma_so_sinh_vien", "student_id", "student_code", "username", "msv"},
    "full_name": {"ho_va_ten", "ho_ten", "full_name", "name", "ten_sinh_vien"},
}
STUDENT_LIST_OPTIONAL_COLUMNS = {
    "gender": {"gioi_tinh", "gender", "sex"},
    "date_of_birth": {"ngay_sinh", "date_of_birth", "dob", "birth_date"},
    "class_name": {"lop", "class", "class_name", "ten_lop"},
    "email": {"email", "mail"},
}


def _student_default_academic_year():
    now = timezone.localtime()
    start_year = now.year if now.month >= 8 else now.year - 1
    return f"{start_year}-{start_year + 1}"


def _student_normalize_header(value):
    value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-zA-Z0-9]+", "_", value.lower()).strip("_")
    return value


def _student_clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _student_parse_birth(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"]:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _student_normalize_gender(value):
    text = _student_clean_cell(value).lower()
    if text in {"nam", "male", "m"}:
        return "Nam"
    if text in {"nu", "nữ", "female", "f"}:
        return "Nữ"
    return _student_clean_cell(value)


def _student_user_is_eligible(user):
    if user is None:
        return False
    if user.is_staff or user.is_superuser:
        return False
    try:
        return not user.userprofile.is_lecturer
    except UserProfile.DoesNotExist:
        return True


def _student_find_user(student_code):
    student_code = normalize_student_code(student_code)
    if not student_code:
        return None

    user = User.objects.select_related("userprofile").filter(username__iexact=student_code).first()
    if _student_user_is_eligible(user):
        return user

    profile = (
        UserProfile.objects.select_related("user_id")
        .filter(student_id__iexact=student_code, is_lecturer=False)
        .first()
    )
    if profile:
        return profile.user_id

    roster_row = (
        StudentRosterStudent.objects.select_related("linked_user_id")
        .filter(student_code__iexact=student_code, linked_user_id__isnull=False)
        .order_by("-created_at", "-pk")
        .first()
    )
    if roster_row and _student_user_is_eligible(roster_row.linked_user_id):
        return roster_row.linked_user_id

    return None


def _student_validate_file(uploaded_file):
    extension = os.path.splitext(uploaded_file.name or "")[1].lower()
    if extension not in STUDENT_LIST_ALLOWED_EXTENSIONS:
        raise ValidationError("Hệ thống chỉ chấp nhận file Excel (.xlsx, .xls) hoặc CSV.")
    if uploaded_file.size > STUDENT_LIST_MAX_FILE_SIZE:
        raise ValidationError("Dung lượng file vượt quá 10MB. Vui lòng kiểm tra lại.")


def _student_resolve_columns(headers):
    mapping = {}
    normalized = [_student_normalize_header(item) for item in headers]

    for target, aliases in STUDENT_LIST_REQUIRED_COLUMNS.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[target] = index
                break
        if target not in mapping:
            raise ValidationError(f"Thiếu cột bắt buộc cho danh sách sinh viên: {target}.")

    for target, aliases in STUDENT_LIST_OPTIONAL_COLUMNS.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                mapping[target] = index
                break

    return mapping


def _student_parse_csv_rows(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if any(str(cell).strip() for cell in row)]


def _student_parse_xlsx_rows(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    worksheet = workbook.active
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        values = ["" if value is None else value for value in row]
        if any(str(cell).strip() for cell in values):
            rows.append(values)
    return rows


def _student_parse_xls_rows(file_bytes):
    try:
        import pandas as pd
    except Exception as exc:
        raise ValidationError("Muốn đọc file .xls, vui lòng cài thêm xlrd==2.0.1 và pandas.") from exc

    try:
        dataframe = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    except Exception as exc:
        raise ValidationError(f"Không thể đọc file .xls: {exc}") from exc

    rows = [list(dataframe.columns)]
    rows.extend(dataframe.fillna("").values.tolist())
    return rows


def _student_parse_records(file_name, file_bytes):
    extension = os.path.splitext(file_name or "")[1].lower()
    if extension == ".csv":
        rows = _student_parse_csv_rows(file_bytes)
    elif extension == ".xlsx":
        rows = _student_parse_xlsx_rows(file_bytes)
    elif extension == ".xls":
        rows = _student_parse_xls_rows(file_bytes)
    else:
        raise ValidationError("Định dạng file không hợp lệ.")

    if len(rows) < 2:
        raise ValidationError("File được chọn không có dữ liệu.")

    headers = rows[0]
    column_map = _student_resolve_columns(headers)

    records = []
    seen_codes = set()
    skipped_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        student_code = normalize_student_code(
            row[column_map["student_code"]] if column_map["student_code"] < len(row) else ""
        )
        full_name = _student_clean_cell(
            row[column_map["full_name"]] if column_map["full_name"] < len(row) else ""
        )

        if not student_code or not full_name:
            skipped_rows.append({"row_number": row_number, "message": f"Dòng {row_number}: thiếu MSSV hoặc họ tên."})
            continue

        if student_code in seen_codes:
            skipped_rows.append(
                {"row_number": row_number, "message": f"Dòng {row_number}: trùng MSSV {student_code} trong cùng file."}
            )
            continue
        seen_codes.add(student_code)

        gender = _student_normalize_gender(
            row[column_map["gender"]]
            if column_map.get("gender") is not None and column_map["gender"] < len(row)
            else ""
        )
        date_of_birth = _student_parse_birth(
            row[column_map["date_of_birth"]]
            if column_map.get("date_of_birth") is not None and column_map["date_of_birth"] < len(row)
            else ""
        )
        class_name = _student_clean_cell(
            row[column_map["class_name"]]
            if column_map.get("class_name") is not None and column_map["class_name"] < len(row)
            else ""
        )
        email = _student_clean_cell(
            row[column_map["email"]]
            if column_map.get("email") is not None and column_map["email"] < len(row)
            else ""
        )

        records.append(
            {
                "student_code": student_code,
                "full_name": full_name,
                "gender": gender,
                "date_of_birth": date_of_birth,
                "class_name": class_name,
                "email": email,
            }
        )

    if not records:
        raise ValidationError("Không tìm thấy sinh viên hợp lệ trong file tải lên.")

    return records, skipped_rows


def _student_normalize_roster_record(record):
    return (
        normalize_student_code(record.get("student_code")),
        (record.get("full_name") or "").strip().lower(),
        (record.get("gender") or "").strip().lower(),
        str(record.get("date_of_birth") or "").strip(),
        (record.get("class_name") or "").strip().lower(),
        (record.get("email") or "").strip().lower(),
    )


def _student_find_exact_duplicate_roster(subject, academic_year, semester, uploaded_file_name, records):
    normalized_file_name = (uploaded_file_name or "").strip()
    incoming_codes = sorted(
        {
            normalize_student_code(record.get("student_code"))
            for record in records
            if normalize_student_code(record.get("student_code"))
        }
    )

    if not normalized_file_name or not incoming_codes:
        return None

    existing_rosters = (
        StudentRosterUpload.objects.filter(
            subject_id=subject,
            academic_year=academic_year,
            semester=semester,
            original_file_name__iexact=normalized_file_name,
        )
        .prefetch_related("students")
        .order_by("-created_at", "-pk")
    )

    for roster in existing_rosters:
        existing_codes = sorted(
            {
                normalize_student_code(student.student_code)
                for student in roster.students.all()
                if normalize_student_code(student.student_code)
            }
        )
        if existing_codes == incoming_codes:
            return roster

    return None

def _ensure_student_enrolled_subject(user, subject):
    if not user or not subject:
        return

    profile, _ = UserProfile.objects.get_or_create(user_id=user)
    if not profile.subjects_enrolled.filter(pk=subject.pk).exists():
        profile.subjects_enrolled.add(subject)


def _revoke_student_enrolled_subject_if_unused(user, subject):
    if not user or not subject:
        return

    profile, _ = UserProfile.objects.get_or_create(user_id=user)

    still_linked = StudentRosterStudent.objects.filter(
        linked_user_id=user,
        student_roster_upload_id__subject_id=subject,
    ).exists()

    if not still_linked:
        profile.subjects_enrolled.remove(subject)


def _sync_roster_enrollments(roster):
    if not roster:
        return 0

    linked_rows = (
        roster.students
        .select_related("linked_user_id")
        .filter(linked_user_id__isnull=False)
    )

    synced_count = 0
    for row in linked_rows:
        profile, _ = UserProfile.objects.get_or_create(user_id=row.linked_user_id)
        before_exists = profile.subjects_enrolled.filter(pk=roster.subject.pk).exists()
        _ensure_student_enrolled_subject(row.linked_user_id, roster.subject)
        after_exists = profile.subjects_enrolled.filter(pk=roster.subject.pk).exists()
        if not before_exists and after_exists:
            synced_count += 1

    return synced_count

def _student_create_roster(subject, created_by, academic_year, semester, uploaded_file, mode="new"):
    if not is_valid_academic_year(academic_year):
        raise ValidationError(ACADEMIC_YEAR_ERROR_MESSAGE)

    if mode == "replace":
        existing_rosters = StudentRosterUpload.objects.filter(
            subject_id=subject,
            academic_year=academic_year,
            semester=semester,
        )

        for old_roster in existing_rosters.prefetch_related("students"):
            linked_users = list(
                old_roster.students.filter(linked_user_id__isnull=False)
                .values_list("linked_user_id", flat=True)
                .distinct()
            )

            if old_roster.uploaded_file:
                old_roster.uploaded_file.delete(save=False)
            old_roster.delete()

            for user_id in linked_users:
                user = User.objects.filter(pk=user_id).first()
                if user:
                    _revoke_student_enrolled_subject_if_unused(user, subject)

    _student_validate_file(uploaded_file)
    file_bytes = uploaded_file.read()
    if not file_bytes:
        raise ValidationError("File được chọn không có dữ liệu.")

    records, skipped_rows = _student_parse_records(uploaded_file.name, file_bytes)

    if mode != "replace":
        duplicate_roster = _student_find_exact_duplicate_roster(
            subject=subject,
            academic_year=academic_year,
            semester=semester,
            uploaded_file_name=uploaded_file.name,
            records=records,
        )
        if duplicate_roster is not None:
            raise ValidationError(
                f"Phát hiện file bị trùng với '{duplicate_roster.original_file_name}' "
                f"(danh sách ID {duplicate_roster.id}) trong cùng môn, năm học và học kỳ."
            )

    existing_codes = set(
        StudentRosterStudent.objects.filter(
            student_roster_upload_id__subject_id=subject,
            student_roster_upload_id__academic_year=academic_year,
            student_roster_upload_id__semester=semester,
        ).values_list("student_code", flat=True)
    )

    duplicate_codes = set()
    for record in records:
        if record["student_code"] in existing_codes:
            duplicate_codes.add(record["student_code"])

    if duplicate_codes:
        skipped_rows.append({
            "row": None,
            "message": f"Các MSSV sau đã tồn tại trong danh sách khác của môn {subject.name} "
                       f"năm học {academic_year} học kỳ {semester}: {', '.join(sorted(duplicate_codes))}. "
                       "Có thể là import trùng hoặc học lại."
        })

    roster = StudentRosterUpload(
        subject_id=subject,
        academic_year=academic_year,
        semester=semester,
        title=os.path.splitext(uploaded_file.name)[0],
        original_file_name=uploaded_file.name,
        total_students=len(records),
        created_by_user_id=created_by,
        status=StudentListUploadStatus.PENDING,
    )
    roster.uploaded_file.save(
        f"{uuid4().hex}_{uploaded_file.name}",
        ContentFile(file_bytes),
        save=False,
    )
    roster.save()

    students = []
    for index, item in enumerate(records, start=1):
        linked_user = _student_find_user(item["student_code"])
        students.append(
            StudentRosterStudent(
                student_roster_upload_id=roster,
                row_number=index,
                student_code=item["student_code"],
                full_name=item["full_name"],
                gender=item["gender"],
                date_of_birth=item["date_of_birth"],
                class_name=item["class_name"],
                email=item["email"],
                linked_user_id=linked_user,
                account_created=bool(linked_user),
                account_status=(
                    StudentRosterAccountStatus.EXISTING
                    if linked_user
                    else StudentRosterAccountStatus.PENDING
                ),
            )
        )

    StudentRosterStudent.objects.bulk_create(students)
    roster.refresh_status()

    # Nếu file roster đã map được tới user có sẵn thì cấp quyền môn ngay
    _sync_roster_enrollments(roster)

    roster.import_warnings = skipped_rows
    return roster


def _student_sync_profile(user, student_row):
    profile, _ = UserProfile.objects.get_or_create(user_id=user)
    profile.is_lecturer = False
    profile.full_name = student_row.full_name
    profile.class_name = student_row.class_name
    profile.student_id = normalize_student_code(student_row.student_code)
    profile.save()
    return profile


def _student_provision_account(student_row, default_password):
    user = student_row.linked_user_id if _student_user_is_eligible(student_row.linked_user_id) else None
    if user is None:
        user = _student_find_user(student_row.student_code)

    normalized_code = normalize_student_code(student_row.student_code) or student_row.student_code
    created = False

    if user is None:
        user = User.objects.create_user(
            username=normalized_code,
            password=default_password,
            email=student_row.email or "",
        )
        created = True
    else:
        user_changed = False
        if student_row.email and user.email != student_row.email:
            user.email = student_row.email
            user_changed = True
        if not user.is_active:
            user.is_active = True
            user_changed = True
        if user_changed:
            user.save()

    _student_sync_profile(user, student_row)
    return user, created, False


def _student_roster_queryset(subject):
    return StudentRosterUpload.objects.filter(subject_id=subject).annotate(
        student_count=Count("students", distinct=True),
        created_accounts_total=Count(
            "students",
            filter=Q(students__account_created=True),
            distinct=True,
        ),
        existing_accounts_total=Count(
            "students",
            filter=Q(students__account_status=StudentRosterAccountStatus.EXISTING),
            distinct=True,
        ),
        new_accounts_total=Count(
            "students",
            filter=Q(students__account_status=StudentRosterAccountStatus.CREATED),
            distinct=True,
        ),
    ).order_by("-created_at", "-pk")


def _student_roster_summary(rosters):
    raw_summary = rosters.aggregate(
        total_files=Count("pk"),
        completed_files=Count(
            "pk",
            filter=Q(status=StudentListUploadStatus.CREATED),
            distinct=True,
        ),
        total_students=Sum("student_count"),
        total_created_accounts=Sum("created_accounts_total"),
    )
    total_files = raw_summary.get("total_files") or 0
    completed_files = raw_summary.get("completed_files") or 0
    total_students = raw_summary.get("total_students") or 0
    total_created_accounts = raw_summary.get("total_created_accounts") or 0

    return {
        "total_files": total_files,
        "completed_files": completed_files,
        "pending_files": max(0, total_files - completed_files),
        "total_students": total_students,
        "total_created_accounts": total_created_accounts,
        "pending_accounts": max(0, total_students - total_created_accounts),
        "all_files_completed": bool(total_files and completed_files == total_files),
    }

@login_required
@lecturer_required
def lecturer_student_list_management(request, subject_code):
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_code=subject_code,
    )

    rosters = _student_roster_queryset(subject)
    keyword = (request.GET.get("q") or "").strip()
    academic_year = (request.GET.get("academic_year") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    status = (request.GET.get("status") or "").strip()

    if keyword:
        rosters = rosters.filter(
            Q(title__icontains=keyword)
            | Q(original_file_name__icontains=keyword)
            | Q(students__student_code__icontains=keyword)
        ).distinct()
    if academic_year and is_valid_academic_year(academic_year):
        rosters = rosters.filter(academic_year=academic_year)
    if semester:
        rosters = rosters.filter(semester=semester)
    if status:
        rosters = rosters.filter(status=status)

    paginator = Paginator(rosters, LIST_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get("page"))
    for roster_item in page_obj.object_list:
        roster_item.safe_academic_year = sanitize_academic_year_for_display(roster_item.academic_year)
    roster_summary = _student_roster_summary(rosters)

    context = {
        "subject": subject,
        "page_obj": page_obj,
        "roster_summary": roster_summary,
        "completed_rosters_preview": rosters.filter(
            status=StudentListUploadStatus.CREATED,
            account_created_at__isnull=False,
        ).order_by("-account_created_at")[:8],
        "filters": {
            "q": keyword,
            "academic_year": academic_year if is_valid_academic_year(academic_year) else "",
            "semester": semester,
            "status": status,
        },
        "academic_year_options": StudentRosterUpload.objects.filter(subject_id=subject)
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year"),
        "semester_options": SemesterChoices.choices,
        "status_options": StudentListUploadStatus.choices,
    }
    context["academic_year_options"] = [
        year for year in context["academic_year_options"] if sanitize_academic_year_for_display(year)
    ]
    return render(request, "qna/lecturer/lecturer_student_list_management.html", context)

@login_required
@lecturer_required
def lecturer_student_list_template(request, subject_code):
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_code=subject_code,
    )

    workbook = OpenpyxlWorkbook()
    worksheet = workbook.active
    worksheet.title = "Danh sách sinh viên"
    worksheet.append(["Mã số sinh viên", "Họ và tên", "Giới tính", "Ngày sinh", "Lớp", "Email"])
    worksheet.append(["SV2024001", "Nguyễn Văn An", "Nam", "12/05/2003", "CNT01", "sv2024001@example.com"])
    worksheet.append(["SV2024002", "Trần Thị Bích", "Nữ", "24/08/2003", "CNT01", "sv2024002@example.com"])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Template_Danh_Sach_SV_{subject.subject_code}.xlsx"'
    workbook.save(response)
    return response


@login_required
@lecturer_required
def lecturer_student_list_detail(request, roster_id):
    roster = get_object_or_404(
        StudentRosterUpload.objects.select_related("subject_id", "created_by_user_id"),
        pk=roster_id,
    )

    if not lecturer_can_access_subject(request.user, roster.subject):
        raise PermissionDenied("Bạn không có quyền truy cập danh sách này.")

    keyword = (request.GET.get("q") or "").strip()
    students_qs = roster.students.select_related("linked_user_id").order_by("row_number")

    if keyword:
        students_qs = students_qs.filter(
            Q(student_code__icontains=keyword)
            | Q(full_name__icontains=keyword)
            | Q(class_name__icontains=keyword)
        )

    paginator = Paginator(students_qs, 8)
    page_obj = paginator.get_page(request.GET.get("page"))

    roster.refresh_status()
    account_breakdown = roster.students.aggregate(
        existing_accounts=Count("pk", filter=Q(account_status=StudentRosterAccountStatus.EXISTING)),
        new_accounts=Count("pk", filter=Q(account_status=StudentRosterAccountStatus.CREATED)),
        pending_accounts=Count("pk", filter=Q(account_status=StudentRosterAccountStatus.PENDING)),
    )
    create_account_modal = request.session.pop("student_account_modal", None) or {}

    context = {
        "roster": roster,
        "subject": roster.subject,
        "page_obj": page_obj,
        "search_query": keyword,
        "pending_accounts_count": roster.pending_accounts_count,
        "created_accounts_count": roster.created_accounts_count,
        "existing_accounts_count": account_breakdown.get("existing_accounts") or 0,
        "new_accounts_count": account_breakdown.get("new_accounts") or 0,
        "uncreated_accounts_count": account_breakdown.get("pending_accounts") or 0,
        "create_account_modal": {
            "open": bool(create_account_modal.get("open")),
            "error": create_account_modal.get("error", ""),
            "existing_students": create_account_modal.get("existing_students", []),
            "need_skip_confirm": bool(create_account_modal.get("need_skip_confirm")),
        },
        "password_rules": [],
    }
    return render(request, "qna/lecturer/lecturer_student_list_detail.html", context)


def _student_account_modal_error(request, roster, message, existing_students=None, need_skip_confirm=False):
    request.session["student_account_modal"] = {
        "open": True,
        "error": message,
        "existing_students": existing_students or [],
        "need_skip_confirm": need_skip_confirm,
    }
    return redirect("qna:lecturer_student_list_detail", roster_id=roster.id)


def _validate_student_batch_password(password):
    if not str(password or "").strip():
        raise ValidationError("Vui lòng nhập mật khẩu.")


@login_required
@require_POST
@lecturer_required
def lecturer_student_list_delete(request, roster_id):
    roster = get_object_or_404(
        StudentRosterUpload.objects.select_related("subject_id"),
        pk=roster_id,
    )

    if not lecturer_can_access_subject(request.user, roster.subject):
        raise PermissionDenied("Bạn không có quyền xóa danh sách này.")

    subject = roster.subject
    subject_code = subject.subject_code

    linked_users = list(
        roster.students.filter(linked_user_id__isnull=False)
        .values_list("linked_user_id", flat=True)
        .distinct()
    )

    if roster.uploaded_file:
        roster.uploaded_file.delete(save=False)
    roster.delete()

    for user_id in linked_users:
        user = User.objects.filter(pk=user_id).first()
        if user:
            _revoke_student_enrolled_subject_if_unused(user, subject)

    messages.success(request, "Xóa danh sách lớp thành công")
    return redirect("qna:lecturer_student_list_management", subject_code=subject_code)


@login_required
@lecturer_required
def lecturer_student_list_upload(request, subject_code):
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_code=subject_code,
    )

    if request.method == "POST":
        academic_year = "" if request.POST.get("academic_year") is None else str(request.POST.get("academic_year"))
        semester = (request.POST.get("semester") or SemesterChoices.HK1).strip()
        if semester not in {choice[0] for choice in SemesterChoices.choices}:
            semester = SemesterChoices.HK1
        mode = request.POST.get("mode", "new")

        files = request.FILES.getlist("files")
        wants_json = _is_ajax_request(request)

        if not is_valid_academic_year(academic_year):
            message = ACADEMIC_YEAR_ERROR_MESSAGE
            if wants_json:
                return JsonResponse({"success": False, "message": message}, status=400)
            messages.error(request, message)
            return redirect("qna:lecturer_student_list_upload", subject_code=subject.subject_code)

        if not files:
            message = "Vui lòng chọn ít nhất 1 file danh sách sinh viên."
            if wants_json:
                return JsonResponse({"success": False, "message": message}, status=400)
            messages.error(request, message)
            return redirect("qna:lecturer_student_list_upload", subject_code=subject.subject_code)

        success_items = []
        errors = []

        for uploaded_file in files:
            try:
                with transaction.atomic():
                    roster = _student_create_roster(
                        subject=subject,
                        created_by=request.user,
                        academic_year=academic_year,
                        semester=semester,
                        uploaded_file=uploaded_file,
                        mode=mode,
                    )
                success_items.append(
                    {
                        "roster_id": roster.id,
                        "file_name": roster.original_file_name,
                        "total_students": roster.total_students,
                        "status": roster.status,
                        "status_label": roster.get_status_display(),
                        "detail_url": f"/lecturer/student-lists/{roster.id}/",
                        "created_at": timezone.localtime(roster.created_at).strftime("%H:%M %d/%m/%Y"),
                        "warnings": getattr(roster, "import_warnings", []),
                    }
                )
            except Exception as exc:
                error_message = " ".join(exc.messages) if isinstance(exc, ValidationError) else str(exc)
                errors.append({"file_name": uploaded_file.name, "message": error_message})

        if wants_json:
            return JsonResponse(
                {
                    "success": bool(success_items) and not errors,
                    "uploaded_count": len(success_items),
                    "results": success_items,
                    "errors": errors,
                    "message": (
                        f"Tải lên thành công {len(success_items)} file danh sách sinh viên."
                        if success_items
                        else "Tải file thất bại."
                    ),
                },
                status=200 if success_items else 400,
            )

        if success_items:
            messages.success(request, f"Tải lên thành công {len(success_items)} file danh sách sinh viên.")
            for item in success_items:
                warnings = item.get("warnings") or []
                if warnings:
                    preview = "; ".join(warning["message"] for warning in warnings[:5])
                    extra_count = max(0, len(warnings) - 5)
                    extra_suffix = f" (+{extra_count} dòng khác)" if extra_count else ""
                    messages.warning(
                        request,
                        f"{item['file_name']}: bỏ qua {len(warnings)} dòng không hợp lệ hoặc trùng lặp. {preview}{extra_suffix}",
                    )

        for error in errors:
            messages.error(request, f"{error['file_name']}: {error['message']}")

        if success_items:
            return redirect("qna:lecturer_student_list_management", subject_code=subject.subject_code)

    context = {
        "subject": subject,
        "default_academic_year": _student_default_academic_year(),
        "semester_options": SemesterChoices.choices,
        "max_upload_mb": STUDENT_LIST_MAX_FILE_SIZE // (1024 * 1024),
    }
    return render(request, "qna/lecturer/lecturer_student_list_upload.html", context)


@login_required
@require_POST
@lecturer_required
def lecturer_student_list_create_accounts(request, roster_id):
    roster = get_object_or_404(
        StudentRosterUpload.objects.select_related("subject_id"),
        pk=roster_id,
    )

    if not lecturer_can_access_subject(request.user, roster.subject):
        raise PermissionDenied("Bạn không có quyền thao tác danh sách này.")

    default_password = (request.POST.get("default_password") or "").strip()
    confirm_password = (request.POST.get("confirm_password") or "").strip()
    skip_existing = str(request.POST.get("skip_existing") or "").strip() in {"1", "true", "yes", "on"}

    if not default_password:
        return _student_account_modal_error(request, roster, "Vui lòng nhập mật khẩu.")

    if default_password != confirm_password:
        return _student_account_modal_error(request, roster, "Mật khẩu xác nhận không khớp.")

    try:
        _validate_student_batch_password(default_password)
    except ValidationError as exc:
        return _student_account_modal_error(request, roster, " ".join(exc.messages))

    roster.refresh_status()
    if roster.pending_accounts_count == 0:
        messages.warning(request, "Danh sách này đã được tạo tài khoản")
        return redirect("qna:lecturer_student_list_detail", roster_id=roster.id)

    existing_rows = []
    pending_rows = []

    def _row_is_created_in_roster(student_row):
        return (
            student_row.account_created
            and student_row.account_status == StudentRosterAccountStatus.CREATED
        )

    def _row_has_canonical_existing_account(student_row, existing_user):
        if not student_row.account_created or not existing_user:
            return False
        return normalize_student_code(existing_user.username) == normalize_student_code(student_row.student_code)

    for student_row in roster.students.select_related("linked_user_id").order_by("row_number"):
        if _row_is_created_in_roster(student_row):
            continue
        existing_user = student_row.linked_user_id or _student_find_user(student_row.student_code)
        if existing_user or student_row.account_status == StudentRosterAccountStatus.EXISTING:
            if _row_has_canonical_existing_account(student_row, existing_user):
                continue
            existing_rows.append(
                {
                    "student_code": student_row.student_code,
                    "full_name": student_row.full_name,
                    "class_name": student_row.class_name,
                }
            )
        else:
            pending_rows.append(student_row)

    # Neu chi co existing_rows (khong co pending), tu dong enroll khong can confirm
    if existing_rows and not pending_rows and not skip_existing:
        skip_existing = True

    if existing_rows and pending_rows and not skip_existing:
        return _student_account_modal_error(
            request,
            roster,
            "Một số sinh viên đã có tài khoản. Vui lòng bấm 'Bỏ qua' để tiếp tục tạo cho các sinh viên còn lại.",
            existing_students=existing_rows,
            need_skip_confirm=True,
        )

    created_count = 0
    skipped_existing_count = len(existing_rows)

    try:
        with transaction.atomic():
            for student_row in roster.students.select_related("linked_user_id").order_by("row_number"):
                if _row_is_created_in_roster(student_row):
                    continue

                existing_user = student_row.linked_user_id or _student_find_user(student_row.student_code)
                if existing_user:
                    student_row.linked_user_id = existing_user
                    student_row.account_created = True
                    student_row.account_status = StudentRosterAccountStatus.EXISTING
                    student_row.save(update_fields=["linked_user_id", "account_created", "account_status"])
                    _ensure_student_enrolled_subject(existing_user, roster.subject)

            for student_row in pending_rows:
                user, created, _ = _student_provision_account(student_row, default_password)
                student_row.linked_user_id = user
                student_row.account_created = True
                student_row.account_status = (
                    StudentRosterAccountStatus.CREATED if created else StudentRosterAccountStatus.EXISTING
                )
                student_row.save(update_fields=["linked_user_id", "account_created", "account_status"])

                # Dù là newly created hay do race-condition thành existing, đều phải cấp quyền môn
                _ensure_student_enrolled_subject(user, roster.subject)

                if created:
                    created_count += 1

            # Dong bo trang thai tat ca roster cung subject (cu + moi cua cung file)
            related_rosters = StudentRosterUpload.objects.filter(
                subject_id=roster.subject_id,
                original_file_name=roster.original_file_name,
            )
            for related_roster in related_rosters:
                related_roster.refresh_status()

    except Exception:
        logger.exception("Batch create student accounts failed for roster_id=%s", roster.id)
        messages.error(
            request,
            "Lỗi đường truyền khi tạo tài khoản. Hệ thống đã hoàn tác toàn bộ, vui lòng tạo lại.",
        )
        return redirect("qna:lecturer_student_list_detail", roster_id=roster.id)

    if skipped_existing_count:
        request.session["student_account_modal"] = {
            "open": True,
            "error": "",
            "existing_students": existing_rows,
            "need_skip_confirm": False,
        }

    messages.success(
        request,
        f"Tạo tài khoản hoàn tất: {created_count} tài khoản mới, {skipped_existing_count} sinh viên đã có sẵn tài khoản.",
    )
    return redirect("qna:lecturer_student_list_detail", roster_id=roster.id)

