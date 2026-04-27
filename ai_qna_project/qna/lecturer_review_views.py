# -*- coding: utf-8 -*-
from __future__ import annotations

import openpyxl
from base64 import b64encode
from datetime import datetime, timedelta
from types import SimpleNamespace
from typing import Any

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Count, Max, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill

from .academic_year import (
    is_valid_academic_year,
    sanitize_academic_year_for_display,
)
from .exam_runtime import LECTURER_HIDDEN_VIOLATION_TYPES
from .lecturer_student_list_views import _student_user_is_eligible, _student_find_user
from .models import (
    DifficultyLevel,
    ExamCode,
    ExamCodeQuestion,
    ExamResult,
    ExamSession,
    ExamSessionGroup,
    ExamSessionRoom,
    ExamSet,
    Question,
    SemesterChoices,
    StudentRosterStudent,
    Subject,
)
from .permissions import (
    lecturer_required,
    get_lecturer_subject_or_404,
    user_can_access_exam_session,
)
from .lecturer_student_accounts import normalize_student_code
User = get_user_model()

from .exam_result_helpers import (
    _with_lecturer_visible_violation_count,
    _has_lecturer_visible_violation,
    _resolve_exam_session_status,
    _safe_float,
    _attach_session_score_meta,
    _build_session_question_details,
    _compute_scores,
    _describe_violation_type,
    _attach_session_duration_meta,
    _attach_session_assignment_meta,
    _build_image_data_url,
    _is_session_finalized,
)
TEN_POINT_SCALE = 10.0
LIST_PAGE_SIZE = 8

# =========================================================
# REVIEW / EXPORT
# =========================================================

def _attach_review_status_meta(session: ExamSession) -> None:
    """
    Chỉ gán các field tạm thời KHÔNG đụng vào property.
    """
    resolved_status = _resolve_exam_session_status(session)
    session.review_status = resolved_status

    if resolved_status == "COMPLETED":
        session.review_status_label = "Đã hoàn thành"
        session.review_status_class = "status-completed"
        session.can_view_detail = True
        session.can_display_score = True


    elif resolved_status == "ABSENT":

        session.review_status_label = "Vắng thi"

        session.review_status_class = "status-absent"

        # Cho phép bấm xem chi tiết bài thi vắng thi

        session.can_view_detail = bool(getattr(session, "id", None))

        # Vắng thi vẫn hiển thị điểm 0/10

        session.can_display_score = True

        if getattr(session, "final_score", None) is None:
            session.final_score = 0

    else:
        session.review_status_label = "Đang thực hiện"
        session.review_status_class = "status-in-progress"
        session.can_view_detail = False
        session.can_display_score = False

    session.status_label = session.review_status_label
    session.status_class = session.review_status_class


def _attach_history_status_meta(session: ExamSession) -> None:
    resolved_status = _resolve_exam_session_status(session, allow_in_progress=True)
    session.history_status = resolved_status or "IN_PROGRESS"

    if session.history_status == "ABSENT":
        session.history_status_label = "Vắng thi"
        session.history_status_class = "status-absent"
    elif session.history_status == "IN_PROGRESS":
        session.history_status_label = "Đang thực hiện"
        session.history_status_class = "status-in-progress"
    else:
        session.history_status_label = "Đã hoàn thành"
        session.history_status_class = "status-completed"


def _apply_review_status(row):
    row.has_violation_warning = bool(getattr(row, "visible_violation_count", 0))
    _attach_review_status_meta(row)


def _can_view_review_detail(row):
    _attach_review_status_meta(row)
    return bool(getattr(row, "can_view_detail", False))

def _resolve_review_student_user(student_row, student_identifier):
    raw_user = (
        student_row.get("linked_user_id")
        or student_row.get("user_id")
        or student_row.get("user")
    )

    if getattr(raw_user, "pk", None):
        return raw_user if _student_user_is_eligible(raw_user) else None

    if raw_user:
        user = User.objects.filter(pk=raw_user).first()
        if _student_user_is_eligible(user):
            return user

    return _student_find_user(student_identifier)

def _get_student_exam_session(user, subject, exam_group):
    if not exam_group:
        return None

    return (
        ExamSession.objects.filter(
            user_id=user,
            subject_id=subject,
            exam_session_group_id=exam_group,
        )
        .order_by("-created_at", "-pk")
        .first()
    )
def _ensure_absent_exam_session(exam_group, student_row, student_identifier):
    if not exam_group or exam_group.computed_status not in {"COMPLETED", "CANCELLED"}:
        return None

    student_user = _resolve_review_student_user(student_row, student_identifier)
    if not student_user:
        return None

    session = _get_student_exam_session(student_user, exam_group.subject_id, exam_group)
    if session:
        return session

    absent_session = ExamSession.objects.create(
        user_id=student_user,
        subject_id=exam_group.subject_id,
        exam_session_group_id=exam_group,
        session_status="PENDING",
        final_score=0,
    )

    if getattr(exam_group, "end_at", None):
        ExamSession.objects.filter(pk=absent_session.pk).update(created_at=exam_group.end_at)
        absent_session.created_at = exam_group.end_at

    return absent_session

def _build_missing_review_rows(exam_groups, existing_session_keys):
    from .lecturer_session_management import _build_exam_group_student_rows

    rows = []

    for exam_group in exam_groups:
        if exam_group.computed_status not in {"COMPLETED", "CANCELLED"}:
            continue

        for student_row in _build_exam_group_student_rows(exam_group):
            student_identifier = student_row.get("student_code") or "-"
            student_key = normalize_student_code(student_identifier)

            if student_key and (exam_group.pk, student_key) in existing_session_keys:
                continue

            absent_session = _ensure_absent_exam_session(
                exam_group,
                student_row,
                student_identifier,
            )

            if absent_session:
                absent_session.student_identifier = student_identifier
                absent_session.student_name = student_row.get("full_name") or student_identifier
                absent_session.student_class_name = student_row.get("class_name") or "-"
                absent_session.exam_date_display = exam_group.exam_date
                absent_session.visible_violation_count = 0
                absent_session.has_violation_warning = False

                # Dùng end_at để sort đúng theo ca thi, không phụ thuộc thời điểm tạo record vắng
                absent_session.created_at = exam_group.end_at or absent_session.created_at

                _attach_session_score_meta(absent_session)
                _attach_review_status_meta(absent_session)
                _attach_session_duration_meta(absent_session)

                rows.append(absent_session)

                if student_key:
                    existing_session_keys.add((exam_group.pk, student_key))

                continue

            # Trường hợp sinh viên chưa có tài khoản/link user thì vẫn hiển thị vắng thi,
            # nhưng không thể bấm xem chi tiết vì không có ExamSession thật.
            row = SimpleNamespace(
                id=None,
                exam_group=exam_group,
                session_status="PENDING",
                completed_at=None,
                cheating_flag=False,
                student_identifier=student_identifier,
                student_name=student_row.get("full_name") or student_identifier,
                student_class_name=student_row.get("class_name") or "-",
                exam_date_display=exam_group.exam_date,
                final_score=0,
                final_display_score=0,
                final_display_score_suffix="/10",
                score_scale_suffix="/10",
                created_at=exam_group.end_at,
                visible_violation_count=0,
            )

            _apply_review_status(row)
            row.can_view_detail = False
            row.can_display_score = True
            rows.append(row)

    return rows


@login_required
@lecturer_required
def lecturer_student_review_screen(request):

    subjects = list(request.user.userprofile.subjects_taught.all().order_by("name"))
    selected_subject_id = (request.GET.get("subject_id") or "").strip()
    raw_exam_group_id = (request.GET.get("exam_group_id") or "").strip()
    student_filter = (request.GET.get("student") or "").strip()

    raw_academic_year = (request.GET.get("academic_year") or "").strip()
    selected_academic_year = raw_academic_year if is_valid_academic_year(raw_academic_year) else ""

    raw_semester = (request.GET.get("semester") or "").strip()
    valid_semesters = {choice[0] for choice in SemesterChoices.choices}
    selected_semester = raw_semester if raw_semester in valid_semesters else ""

    raw_status = (request.GET.get("status") or "").strip()
    valid_review_statuses = {"COMPLETED", "ABSENT", "IN_PROGRESS"}
    selected_status = raw_status if raw_status in valid_review_statuses else ""

    selected_exam_group_id = int(raw_exam_group_id) if raw_exam_group_id.isdigit() else None
    selected_subject = next((subject for subject in subjects if str(subject.pk) == selected_subject_id), None)

    exam_groups = []
    year_options = []
    review_rows = []
    page_obj = None

    review_status_choices = [
        ("", "Trạng thái: Tất cả"),
        ("COMPLETED", "Đã hoàn thành"),
        ("ABSENT", "Vắng thi"),
        ("IN_PROGRESS", "Đang thực hiện"),
    ]

    if selected_subject:
        all_exam_groups_qs = ExamSessionGroup.objects.filter(subject_id=selected_subject)

        year_options = [
            year for year in all_exam_groups_qs
            .values_list("academic_year", flat=True)
            .distinct()
            .order_by("-academic_year")
            if sanitize_academic_year_for_display(year)
        ]

        exam_groups_qs = all_exam_groups_qs

        if selected_academic_year:
            exam_groups_qs = exam_groups_qs.filter(academic_year=selected_academic_year)

        if selected_semester:
            exam_groups_qs = exam_groups_qs.filter(semester=selected_semester)

        exam_groups = list(exam_groups_qs.order_by("-exam_date", "-pk"))
        sessions = _with_lecturer_visible_violation_count(
            ExamSession.objects.filter(
                subject_id=selected_subject,
                exam_session_group_id__isnull=False,
            ).select_related(
                "user_id",
                "user_id__userprofile",
                "exam_session_group_id",
            )
        )
        if selected_academic_year:
            sessions = sessions.filter(exam_session_group_id__academic_year=selected_academic_year)

        if selected_semester:
            sessions = sessions.filter(exam_session_group_id__semester=selected_semester)

        if selected_exam_group_id:
            sessions = sessions.filter(exam_session_group_id=selected_exam_group_id)

        selected_exam_groups = [
            exam_group for exam_group in exam_groups if exam_group.pk == selected_exam_group_id
        ] if selected_exam_group_id else exam_groups
        existing_session_keys = set()
        sessions = sessions.order_by("-created_at", "-pk")

        for session in sessions:
            main_avg, final_total = _compute_scores(session)
            session.main_avg = main_avg
            if session.final_score != final_total:
                session.final_score = final_total

            _attach_session_score_meta(session)

            profile = getattr(session.user_id, "userprofile", None)

            linked_roster_row = (
                StudentRosterStudent.objects
                .filter(linked_user_id=session.user_id, student_roster_upload_id__subject_id=session.subject_id)
                .order_by("-student_roster_student_id")
                .first()
            )

            session.student_identifier = (
                    (linked_roster_row.student_code if linked_roster_row else None)
                    or getattr(profile, "student_id", None)
                    or session.user_id.username
            )

            session.student_name = (
                    (linked_roster_row.full_name if linked_roster_row else None)
                    or getattr(profile, "full_name", None)
                    or session.user_id.username
            )

            session.student_class_name = (
                    (linked_roster_row.class_name if linked_roster_row else None)
                    or getattr(profile, "class_name", None)
                    or "-"
            )
            session.exam_date_display = (
                session.exam_group.exam_date
                if session.exam_group and session.exam_group.exam_date
                else session.completed_at.date() if session.completed_at
                else session.created_at.date()
            )
            session.has_violation_warning = bool(getattr(session, "visible_violation_count", 0))
            _attach_review_status_meta(session)
            _attach_session_duration_meta(session)
            review_rows.append(session)

            if session.exam_group:
                session_key = normalize_student_code(session.student_identifier)
                if session_key:
                    existing_session_keys.add((session.exam_group.pk, session_key))

        review_rows.extend(_build_missing_review_rows(selected_exam_groups, existing_session_keys))

        if student_filter:
            keyword = student_filter.lower()
            review_rows = [
                row for row in review_rows
                if keyword in (row.student_identifier or "").lower()
                or keyword in (row.student_name or "").lower()
                or keyword in (row.student_class_name or "").lower()
                or keyword in ((row.exam_group.group_name if getattr(row, "exam_group", None) else "") or "").lower()
            ]
        if selected_status:
            review_rows = [
                row for row in review_rows
                if getattr(row, "review_status", None) == selected_status
            ]

        default_sort_time = timezone.make_aware(datetime(1970, 1, 1))
        review_rows.sort(
            key=lambda item: getattr(item, "created_at", None) or default_sort_time,
            reverse=True,
        )

    completed_count = 0
    absent_count = 0
    total_score = 0.0
    flagged_count = 0

    for row in review_rows:
        if getattr(row, "has_violation_warning", False):
            flagged_count += 1

        status_label = getattr(row, "review_status_label", "")
        if status_label == "Đã hoàn thành":
            completed_count += 1
            total_score += float(getattr(row, "final_score", 0.0) or 0.0)
        elif status_label == "Vắng thi":
            absent_count += 1

    average_score = round(total_score / completed_count, 1) if completed_count else 0

    sessions_page = []
    if selected_subject:
        paginator = Paginator(review_rows, LIST_PAGE_SIZE)
        page_obj = paginator.get_page(request.GET.get("page"))
        sessions_page = page_obj.object_list

    return render(
        request,
        "qna/lecturer/lecturer_student_review.html",
        {
            "subjects": subjects,
            "selected_subject": selected_subject,
            "subject": selected_subject,
            "exam_groups": exam_groups,
            "selected_exam_group_id": selected_exam_group_id,
            "sessions": sessions_page,
            "page_obj": page_obj,
            "student_filter": student_filter,
            "average_score": average_score,
            "completed_count": completed_count,
            "absent_count": absent_count,
            "flagged_count": flagged_count,
            "year_options": year_options,
            "semester_choices": SemesterChoices.choices,
            "review_status_choices": review_status_choices,
            "filters": {
                "subject_id": selected_subject_id,
                "exam_group_id": raw_exam_group_id,
                "student": student_filter,
                "academic_year": selected_academic_year,
                "semester": selected_semester,
                "status": selected_status,
            },
        },
    )


@login_required
@lecturer_required
def lecturer_student_review(request, subject_code):
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_code=subject_code,
    )
    query = request.GET.copy()
    query["subject_id"] = str(subject.id)
    return redirect(f"{reverse('qna:lecturer_student_review_screen')}?{query.urlencode()}")


@login_required
@lecturer_required
def lecturer_session_detail(request, session_id):

    session = get_object_or_404(
        ExamSession.objects.select_related(
            "subject_id",
            "user_id",
            "user_id__userprofile",
            "exam_session_group_id",
        ),
        pk=session_id,
    )
    if not user_can_access_exam_session(request.user, session):
        raise PermissionDenied("Bạn không có quyền")
    main_avg, final_total = _compute_scores(session)
    if session.final_score != final_total:
        session.final_score = final_total
    _attach_review_status_meta(session)
    if getattr(session, "review_status", None) == "IN_PROGRESS":
        messages.error(request, "Phiên thi đang diễn ra. Chưa thể xem chi tiết.")
        return redirect("qna:lecturer_student_review_screen")
    _attach_session_score_meta(session)
    _attach_session_duration_meta(session)
    _attach_session_assignment_meta(session)
    question_details = _build_session_question_details(session)

    session.detail_status_label = session.review_status_label
    session.detail_status_class = session.review_status_class

    violation_images = [
        {
            "image_data_url": image.get_image_data_url,
            "timestamp": image.timestamp,
            "violation_description": _describe_violation_type(image.violation_type),
        }
        for image in session.violation_images.exclude(
            violation_type__in=LECTURER_HIDDEN_VIOLATION_TYPES
        ).order_by("-timestamp")
    ]

    face_image_url = _build_image_data_url(
        session.face_image_blob,
        session.face_image_mime,
    )
    id_card_image_url = _build_image_data_url(
        session.id_card_image_blob,
        session.id_card_image_mime,
    )

    return render(
        request,
        "qna/lecturer/lecturer_session_detail.html",
        {
            "session": session,
            "question_details": question_details,
            "main_avg": main_avg,
            "final_total": final_total,
            "violation_images": violation_images,
            "face_image_url": face_image_url,
            "id_card_image_url": id_card_image_url,
        },
    )


@login_required
def lecturer_export_reports_screen(request):
    return redirect("qna:lecturer_student_review_screen")


@login_required
@lecturer_required
def lecturer_export_exam_results_screen(request):
    subject_id = request.GET.get("subject_id")
    exam_group_id = request.GET.get("exam_group_id")
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_id=subject_id,
    )

    if exam_group_id:
        exam_group = get_object_or_404(ExamSessionGroup, pk=exam_group_id, subject_id=subject)
        sessions = ExamSession.objects.filter(subject_id=subject, exam_session_group_id=exam_group).select_related(
            "user_id",
            "user_id__userprofile",
        )
    else:
        sessions = ExamSession.objects.filter(subject_id=subject).select_related("user_id", "user_id__userprofile")

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả thi"

    # YÊU CẦU 1: Loại bỏ cột điểm phụ
    headers = ["STT", "Mã SV", "Họ tên", "Lớp", "Ngày thi", "Điểm tổng", "Cảnh báo gian lận"]
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, session in enumerate(sessions, 1):
        main_avg, final_total = _compute_scores(session)
        profile = session.user.userprofile
        ws.append(
            [
                idx,
                session.user.username,
                profile.full_name,
                profile.class_name,
                session.created_at.strftime("%d/%m/%Y %H:%M"),
                f"{final_total:.2f}",
                "CÓ" if _has_lecturer_visible_violation(session) else "KHÔNG",
            ]
        )

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Ket_qua_thi_{subject.subject_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@lecturer_required
def lecturer_export_exam_results(request, subject_code):
    subject = get_lecturer_subject_or_404(
        request.user,
        subject_code=subject_code,
    )

    sessions = (
        ExamSession.objects.filter(subject_id=subject, session_status="COMPLETED")
        .select_related("user_id", "exam_session_group_id")
        .order_by("-created_at")
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ket_qua_{subject.subject_code}.xlsx"'

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kết quả thi"

    headers = [
        "STT",
        "Họ tên",
        "Username",
        "Mã sinh viên",
        "Lớp",
        "Môn học",
        "Mã môn",
        "Ca thi",
        "Ngày thi",
        "Điểm cuối",
        "Trạng thái hoàn thành",
        "Xác thực khuôn mặt",
        "Cảnh báo gian lận"
    ]
    ws.append(headers)

    for idx, session in enumerate(sessions, start=1):
        profile = getattr(session.user, "userprofile", None)
        ws.append(
            [
                idx,
                profile.full_name if profile and profile.full_name else session.user.get_full_name() or session.user.username,
                session.user.username,
                profile.student_id if profile else "",
                profile.class_name if profile else "",
                subject.name,
                subject.subject_code,
                session.exam_group.group_name if session.exam_group else "",
                session.created_at.strftime("%d/%m/%Y %H:%M") if session.created_at else "",
                session.final_score if session.final_score is not None else "",
                "Đã hoàn thành" if _is_session_finalized(session) else "Chưa hoàn thành",
                session.get_verification_status_display() if hasattr(session,
                                                                     "get_verification_status_display") else session.verification_status,
                "CÓ" if _has_lecturer_visible_violation(session) else "KHÔNG"
            ]
        )

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(response)
    return response
